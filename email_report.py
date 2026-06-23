
import smtplib

import openpyxl

from email.mime.multipart import MIMEMultipart

from email.mime.base import MIMEBase

from email.mime.text import MIMEText

from email import encoders

from datetime import datetime

import os



SENDER_EMAIL    = "attendencesystem15@gmail.com"

SENDER_PASSWORD = "oafj oksk wvhy qket"

FACULTY_EMAIL   = "jaiswalbhoomi1911@gmail.com"

SUBJECT_NAME    = "Computer Science 101"



def generate_excel(attendance_data, subject, date_str):

    filename = f"attendance_{subject.replace(' ', '_')}_{date_str}.xlsx"

    filepath = os.path.join(os.path.dirname(__file__), filename)

    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Attendance"

    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill  = PatternFill("solid", fgColor="1F4E79")

    present_fill = PatternFill("solid", fgColor="C6EFCE")

    absent_fill  = PatternFill("solid", fgColor="FFC7CE")

    headers = ["#","Student Name","Roll No","Date","Snap 1","Snap 2","Snap 3","Snap 4","Snap 5","Snaps Present","Total Snaps","Verdict"]

    ws.append(headers)

    for cell in ws[1]:

        cell.font      = Font(bold=True, color="FFFFFF")

        cell.fill      = header_fill

        cell.alignment = Alignment(horizontal="center")

    for i, student in enumerate(attendance_data, start=1):

        snaps         = student["snaps"]

        snaps_present = sum(snaps)

        total_snaps   = len(snaps)

        snap_symbols  = ["✓" if s else "✗" for s in snaps]

        while len(snap_symbols) < 5:

            snap_symbols.append("—")

        row = [i, student["name"], student["roll"], date_str,

               snap_symbols[0], snap_symbols[1], snap_symbols[2],

               snap_symbols[3], snap_symbols[4],

               snaps_present, total_snaps, student["verdict"]]

        ws.append(row)

        verdict_cell = ws.cell(row=i+1, column=12)

        if student["verdict"] == "Present":

            verdict_cell.fill = present_fill

            verdict_cell.font = Font(bold=True, color="276221")

        else:

            verdict_cell.fill = absent_fill

            verdict_cell.font = Font(bold=True, color="9C0006")

        for col in range(5, 10):

            ws.cell(row=i+1, column=col).alignment = Alignment(horizontal="center")

    ws.append([])

    total   = len(attendance_data)

    present = sum(1 for s in attendance_data if s["verdict"] == "Present")

    absent  = total - present

    pct     = round((present / total * 100), 1) if total > 0 else 0

    ws.append(["","Total Students","","","","","","","","","",total])

    ws.append(["","Present","","","","","","","","","",present])

    ws.append(["","Absent","","","","","","","","","",absent])

    ws.append(["","Attendance %","","","","","","","","","",f"{pct}%"])

    ws.column_dimensions["A"].width = 5

    ws.column_dimensions["B"].width = 22

    ws.column_dimensions["C"].width = 12

    ws.column_dimensions["D"].width = 14

    for col in ["E","F","G","H","I"]:

        ws.column_dimensions[col].width = 8

    ws.column_dimensions["J"].width = 14

    ws.column_dimensions["K"].width = 12

    ws.column_dimensions["L"].width = 12

    wb.save(filepath)

    print(f"Excel saved: {filename}")

    return filepath



def send_email(filepath, subject, date_str, stats):

    present = stats["present"]

    absent  = stats["absent"]

    total   = stats["total"]

    pct     = round((present / total * 100), 1) if total > 0 else 0

    msg = MIMEMultipart()

    msg["From"]    = SENDER_EMAIL

    msg["To"]      = FACULTY_EMAIL

    msg["Subject"] = f"Attendance Report - {subject} - {date_str}"

    body = f"""Dear Faculty,



The attendance report for today's class is ready.



Subject   : {subject}

Date      : {date_str}

Present   : {present} / {total}

Absent    : {absent} / {total}

Attendance: {pct}%



Please find the detailed Excel sheet attached.

Attendance was recorded automatically via the Face Recognition System.



Regards,

Automated Attendance System"""

    msg.attach(MIMEText(body, "plain"))

    with open(filepath, "rb") as f:

        part = MIMEBase("application", "octet-stream")

        part.set_payload(f.read())

    encoders.encode_base64(part)

    part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(filepath)}")

    msg.attach(part)

    try:

        print(f"Sending email to {FACULTY_EMAIL}...")

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:

            server.login(SENDER_EMAIL, SENDER_PASSWORD)

            server.sendmail(SENDER_EMAIL, FACULTY_EMAIL, msg.as_string())

        print("Email sent successfully!")

    except Exception as e:

        print(f"Email failed: {e}")



def send_attendance_report(attendance_data, subject=SUBJECT_NAME):

    date_str = datetime.now().strftime("%d-%b-%Y")

    filepath = generate_excel(attendance_data, subject, date_str)

    total    = len(attendance_data)

    present  = sum(1 for s in attendance_data if s["verdict"] == "Present")

    absent   = total - present

    stats    = {"total": total, "present": present, "absent": absent}

    send_email(filepath, subject, date_str, stats)



if __name__ == "__main__":

    dummy_data = [

        {"name": "Arjun Sharma",  "roll": "CS001", "snaps": [True,  True,  True,  False, True],  "verdict": "Present"},

        {"name": "Priya Nair",    "roll": "CS002", "snaps": [True,  True,  True,  True,  True],  "verdict": "Present"},

        {"name": "Rahul Verma",   "roll": "CS003", "snaps": [False, False, True,  False, False], "verdict": "Absent"},

        {"name": "Divya Menon",   "roll": "CS004", "snaps": [True,  False, True,  True,  False], "verdict": "Present"},

        {"name": "Kiran Patel",   "roll": "CS005", "snaps": [False, False, False, False, False], "verdict": "Absent"},

        {"name": "Sneha Reddy",   "roll": "CS006", "snaps": [True,  True,  False, True,  True],  "verdict": "Present"},

        {"name": "Amit Joshi",    "roll": "CS007", "snaps": [True,  True,  True,  True,  False], "verdict": "Present"},

        {"name": "Lakshmi Iyer",  "roll": "CS008", "snaps": [False, True,  False, False, False], "verdict": "Absent"},

    ]

    send_attendance_report(dummy_data, subject="Computer Science 101")

