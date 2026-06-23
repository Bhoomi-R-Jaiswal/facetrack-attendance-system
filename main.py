import face_recognition
import cv2
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime
import os
import time

SENDER_EMAIL     = "attendencesystem15@gmail.com"
SENDER_PASSWORD  = "oafj oksk wvhy qket"
FACULTY_EMAIL    = "jaiswalbhoomi1911@gmail.com"
SUBJECT_NAME     = "Electronics and Communication"
CLASS_DURATION   = 5
SNAP_INTERVAL    = 1
MIN_SNAPS        = 4
TOLERANCE        = 0.6
CAMERA_INDEX     = 0
STUDENTS_FOLDER  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "students")
SNAPSHOTS_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "snapshots")

os.makedirs(SNAPSHOTS_FOLDER, exist_ok=True)
os.makedirs(STUDENTS_FOLDER,  exist_ok=True)

def enroll_students():
    print("\n" + "="*50)
    print("  STEP 1 - Enrolling students from photos")
    print("="*50)
    known_encodings = []
    known_names     = []
    known_rolls     = []
    files = [f for f in os.listdir(STUDENTS_FOLDER)
             if f.lower().endswith((".jpg", ".jpeg", ".png"))]
    if not files:
        print("  ERROR: No photos found in students/ folder.")
        return [], [], []
    for filename in files:
        base  = os.path.splitext(filename)[0]
        parts = base.split("_")
        if len(parts) < 2:
            print(f"  SKIP: {filename}")
            continue
        roll = parts[0]
        name = " ".join(parts[1:]).replace("-", " ").title()
        image_path = os.path.join(STUDENTS_FOLDER, filename)
        try:
            image     = face_recognition.load_image_file(image_path)
            encodings = face_recognition.face_encodings(image)
        except Exception as e:
            print(f"  ERROR reading {filename}: {e}")
            continue
        if len(encodings) == 0:
            print(f"  WARNING: No face found in {filename} - use a clearer photo")
            continue
        known_encodings.append(encodings[0])
        known_names.append(name)
        known_rolls.append(roll)
        print(f"  Enrolled: [{roll}] {name}")
    print(f"\n  Total enrolled: {len(known_names)} student(s)")
    return known_encodings, known_names, known_rolls

def draw_boxes(frame, face_locations, face_names, face_rolls, confidences):
    for (top, right, bottom, left), name, roll, conf in zip(
            face_locations, face_names, face_rolls, confidences):
        # Scale back up since we shrunk the frame
        top    *= 4
        right  *= 4
        bottom *= 4
        left   *= 4

        if name == "Unknown":
            color = (0, 0, 220)       # red for unknown
            label = "Unknown"
            sublabel = "Not in database"
        else:
            color = (0, 200, 0)       # green for recognised
            label = f"{roll} - {name}"
            sublabel = f"Match: {conf}%"

        # Draw rectangle around face
        cv2.rectangle(frame, (left, top), (right, bottom), color, 2)

        # Draw filled label box at bottom of rectangle
        cv2.rectangle(frame, (left, bottom), (right, bottom + 45), color, cv2.FILLED)

        # Draw name
        cv2.putText(frame, label,
                    (left + 6, bottom + 16),
                    cv2.FONT_HERSHEY_DUPLEX, 0.55, (255, 255, 255), 1)

        # Draw confidence
        cv2.putText(frame, sublabel,
                    (left + 6, bottom + 36),
                    cv2.FONT_HERSHEY_DUPLEX, 0.45, (255, 255, 255), 1)

    return frame

def process_frame(frame, known_encodings, known_names, known_rolls):
    small = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb   = cv2.cvtColor(small, cv2.COLOR_BGR2RGB)

    locations = face_recognition.face_locations(rgb)
    encodings = face_recognition.face_encodings(rgb, locations)

    face_names = []
    face_rolls = []
    confidences = []

    for encoding in encodings:
        matches   = face_recognition.compare_faces(known_encodings, encoding, tolerance=TOLERANCE)
        distances = face_recognition.face_distance(known_encodings, encoding)

        name = "Unknown"
        roll = ""
        conf = 0

        if True in matches:
            best = np.argmin(distances)
            if matches[best]:
                name = known_names[best]
                roll = known_rolls[best]
                conf = round((1 - distances[best]) * 100, 1)

        face_names.append(name)
        face_rolls.append(roll)
        confidences.append(conf)

    return locations, face_names, face_rolls, confidences

def take_snapshot(frame, snap_num, known_encodings, known_names, known_rolls):
    print(f"\n  Snapshot {snap_num} at {datetime.now().strftime('%H:%M:%S')}")
    snap_filename = f"snap{snap_num}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
    snap_path     = os.path.join(SNAPSHOTS_FOLDER, snap_filename)
    cv2.imwrite(snap_path, frame)
    print(f"  Saved: {snap_filename}")

    locations, face_names, face_rolls, confidences = process_frame(
        frame, known_encodings, known_names, known_rolls)

    found = []
    if not locations:
        print("  No faces detected in snapshot.")
    for name, roll, conf in zip(face_names, face_rolls, confidences):
        if name != "Unknown":
            found.append(name)
            print(f"  Recognised: [{roll}] {name} ({conf}%)")
        else:
            print(f"  Unknown face detected")

    return found

def generate_excel(attendance_data, date_str):
    print("\n" + "="*50)
    print("  STEP 3 - Generating Excel report")
    print("="*50)
    filename = f"attendance_{SUBJECT_NAME.replace(' ','_')}_{date_str}.xlsx"
    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    present_fill = PatternFill("solid", fgColor="C6EFCE")
    absent_fill  = PatternFill("solid", fgColor="FFC7CE")
    summary_fill = PatternFill("solid", fgColor="F2F2F2")
    headers = ["#","Roll No","Student Name","Date","Snap 1","Snap 2","Snap 3","Snap 4","Snap 5","Snaps Present","Total Snaps","Verdict"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    for i, s in enumerate(attendance_data, start=1):
        syms = ["Y" if x else "N" for x in s["snaps"]]
        while len(syms) < 5:
            syms.append("-")
        row = [i, s["roll"], s["name"], date_str,
               syms[0], syms[1], syms[2], syms[3], syms[4],
               sum(s["snaps"]), len(s["snaps"]), s["verdict"]]
        ws.append(row)
        vcell = ws.cell(row=i+1, column=12)
        if s["verdict"] == "Present":
            vcell.fill = present_fill
            vcell.font = Font(bold=True, color="276221")
        else:
            vcell.fill = absent_fill
            vcell.font = Font(bold=True, color="9C0006")
        for col in range(5, 10):
            ws.cell(row=i+1, column=col).alignment = Alignment(horizontal="center")
    total   = len(attendance_data)
    present = sum(1 for s in attendance_data if s["verdict"] == "Present")
    absent  = total - present
    pct     = round(present / total * 100, 1) if total > 0 else 0
    ws.append([])
    summary = [
        ["","","Total Students","","","","","","","","",total],
        ["","","Present","","","","","","","","",present],
        ["","","Absent","","","","","","","","",absent],
        ["","","Attendance %","","","","","","","","",f"{pct}%"],
    ]
    for row in summary:
        ws.append(row)
        r = ws.max_row
        ws.cell(r, 3).font  = Font(bold=True)
        ws.cell(r, 3).fill  = summary_fill
        ws.cell(r, 12).font = Font(bold=True)
        ws.cell(r, 12).fill = summary_fill
    widths = {"A":4,"B":10,"C":22,"D":14,"E":8,"F":8,"G":8,"H":8,"I":8,"J":14,"K":12,"L":12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    wb.save(filepath)
    print(f"  Saved: {filename}")
    return filepath, present, absent, total, pct

def send_email(filepath, date_str, present, absent, total, pct):
    print("\n" + "="*50)
    print("  STEP 4 - Sending email to faculty")
    print("="*50)
    msg = MIMEMultipart()
    msg["From"]    = SENDER_EMAIL
    msg["To"]      = FACULTY_EMAIL
    msg["Subject"] = f"Attendance Report - {SUBJECT_NAME} - {date_str}"
    body = f"""Dear Faculty,

Attendance report for today's class is attached.

Subject    : {SUBJECT_NAME}
Date       : {date_str}
Present    : {present} / {total}
Absent     : {absent} / {total}
Attendance : {pct}%

A student is marked Present if detected in {MIN_SNAPS} or more of the 5 snapshots.
Snapshots taken every {SNAP_INTERVAL} minute(s) during the {CLASS_DURATION}-minute class.

Regards,
Automated Face Recognition Attendance System"""
    msg.attach(MIMEText(body, "plain"))
    with open(filepath, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(filepath)}")
    msg.attach(part)
    try:
        print(f"  Connecting to Gmail...")
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            server.sendmail(SENDER_EMAIL, FACULTY_EMAIL, msg.as_string())
        print(f"  Email sent to {FACULTY_EMAIL}")
    except Exception as e:
        print(f"  Email failed: {e}")

def main():
    print("\n" + "="*50)
    print("   FACE RECOGNITION ATTENDANCE SYSTEM")
    print("="*50)
    print(f"  Subject  : {SUBJECT_NAME}")
    print(f"  Duration : {CLASS_DURATION} minutes")
    print(f"  Snaps    : every {SNAP_INTERVAL} min")
    print(f"  Present if detected in >= {MIN_SNAPS} snaps")

    known_encodings, known_names, known_rolls = enroll_students()
    if not known_names:
        print("\nNo students enrolled. Exiting.")
        return

    tracker = {}
    for name, roll in zip(known_names, known_rolls):
        tracker[name] = {"roll": roll, "snaps": []}

    camera = cv2.VideoCapture(CAMERA_INDEX)
    if not camera.isOpened():
        print(f"\nERROR: Cannot open camera {CAMERA_INDEX}.")
        print("Try changing CAMERA_INDEX to 1.")
        return

    print(f"\n  Camera opened successfully.")
    print(f"  A live window will open — green box = recognised, red box = unknown.")
    print(f"  Taking {CLASS_DURATION // SNAP_INTERVAL} snapshots over {CLASS_DURATION} minutes.")
    print(f"  Press Q in the camera window to stop early.\n")

    total_snaps    = CLASS_DURATION // SNAP_INTERVAL
    snap_count     = 0
    class_start    = time.time()
    last_snap_time = 0
    next_snap_due  = time.time()

    # For live display — process every 3rd frame to keep it smooth
    frame_count    = 0
    last_locations = []
    last_names     = []
    last_rolls     = []
    last_confs     = []

    while snap_count < total_snaps:
        success, frame = camera.read()
        if not success:
            break

        frame_count += 1
        elapsed = time.time() - class_start

        # Process face detection every 3 frames for smooth display
        if frame_count % 3 == 0:
            last_locations, last_names, last_rolls, last_confs = process_frame(
                frame, known_encodings, known_names, known_rolls)

        # Draw boxes on the live frame
        display_frame = draw_boxes(
            frame.copy(), last_locations, last_names, last_rolls, last_confs)

        # Show countdown to next snap
        time_remaining = max(0, next_snap_due - time.time())
        mins = int(time_remaining) // 60
        secs = int(time_remaining) % 60

        # Overlay info on screen
        cv2.rectangle(display_frame, (0, 0), (640, 36), (30, 30, 30), cv2.FILLED)
        cv2.putText(display_frame,
                    f"Snap {snap_count+1}/{total_snaps}  |  Next snap in: {mins}m {secs}s  |  Press Q to quit",
                    (10, 24), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255, 255, 255), 1)

        cv2.imshow("Attendance System - Face Recognition", display_frame)

        # Take snapshot at interval
        # Take snapshot at interval
        now = time.time()
        if now >= next_snap_due and snap_count < total_snaps:
            snap_count    += 1
            next_snap_due  = now + (SNAP_INTERVAL * 60)
            found = take_snapshot(frame, snap_count,
                                  known_encodings, known_names, known_rolls)
            for name in tracker:
                tracker[name]["snaps"].append(name in found)
            print(f"  Progress: {snap_count}/{total_snaps} snaps done")
            if snap_count < total_snaps:
                print(f"  Next snap in {SNAP_INTERVAL} minute(s)...")

        # Press Q to quit
        if cv2.waitKey(1) & 0xFF == ord("q"):
            print("\n  Stopped early by user.")
            break

    camera.release()
    cv2.destroyAllWindows()
    print("\n  Camera closed.")

    # Results
    print("\n" + "="*50)
    print("  STEP 2 - Attendance Results")
    print("="*50)
    attendance_data = []
    for name, data in tracker.items():
        snaps_present = sum(data["snaps"])
        verdict = "Present" if snaps_present >= MIN_SNAPS else "Absent"
        attendance_data.append({
            "name"   : name,
            "roll"   : data["roll"],
            "snaps"  : data["snaps"],
            "verdict": verdict
        })
        status = "PRESENT" if verdict == "Present" else "ABSENT"
        print(f"  {data['roll']:<10} {name:<25} {snaps_present}/{len(data['snaps'])} snaps  {status}")

    date_str = datetime.now().strftime("%d-%b-%Y")
    filepath, present, absent, total, pct = generate_excel(attendance_data, date_str)
    send_email(filepath, date_str, present, absent, total, pct)

    print("\n" + "="*50)
    print("  ALL DONE")
    print(f"  Present : {present}/{total}")
    print(f"  Absent  : {absent}/{total}")
    print(f"  Report emailed to {FACULTY_EMAIL}")
    print("="*50 + "\n")

if __name__ == "__main__":
    main()
