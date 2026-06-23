from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
from flask_socketio import SocketIO, emit
import face_recognition
import cv2
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime
import os
import time
import threading
import base64
import json

app = Flask(__name__)
app.config["SECRET_KEY"] = "attendance123"
socketio = SocketIO(app, cors_allowed_origins="*")

# ── CONFIG ────────────────────────────────────────────────────────
SENDER_EMAIL     = "attendencesystem15@gmail.com"
SENDER_PASSWORD  = "oafj oksk wvhy qket"
FACULTY_EMAIL    = "jaiswalbhoomi1911@gmail.com"
TOLERANCE        = 0.6
CAMERA_INDEX     = 0
BASE_DIR         = os.path.dirname(os.path.abspath(__file__))
STUDENTS_FOLDER  = os.path.join(BASE_DIR, "students")
SNAPSHOTS_FOLDER = os.path.join(BASE_DIR, "snapshots")
REPORTS_FOLDER   = os.path.join(BASE_DIR, "reports")
HISTORY_FILE     = os.path.join(BASE_DIR, "history.json")

for folder in [STUDENTS_FOLDER, SNAPSHOTS_FOLDER, REPORTS_FOLDER]:
    os.makedirs(folder, exist_ok=True)

# ── Global session state ──────────────────────────────────────────
session_state = {
    "running"       : False,
    "subject"       : "",
    "duration"      : 50,
    "interval"      : 10,
    "min_snaps"     : 4,
    "snap_count"    : 0,
    "total_snaps"   : 5,
    "tracker"       : {},
    "log"           : [],
    "start_time"    : None,
    "next_snap_due" : None,
}

known_encodings = []
known_names     = []
known_rolls     = []

# ── Load history ──────────────────────────────────────────────────
def load_history():
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE) as f:
            return json.load(f)
    return []

def save_history(record):
    history = load_history()
    history.append(record)
    with open(HISTORY_FILE, "w") as f:
        json.dump(history, f, indent=2)

# ── Enroll students ───────────────────────────────────────────────
def enroll_students():
    global known_encodings, known_names, known_rolls
    known_encodings = []
    known_names     = []
    known_rolls     = []
    files = [f for f in os.listdir(STUDENTS_FOLDER)
             if f.lower().endswith((".jpg", ".jpeg", ".png"))]
    for filename in files:
        base  = os.path.splitext(filename)[0]
        parts = base.split("_")
        if len(parts) < 2:
            continue
        roll = parts[0]
        name = " ".join(parts[1:]).replace("-", " ").title()
        try:
            image     = face_recognition.load_image_file(
                            os.path.join(STUDENTS_FOLDER, filename))
            encodings = face_recognition.face_encodings(image)
            if encodings:
                known_encodings.append(encodings[0])
                known_names.append(name)
                known_rolls.append(roll)
        except:
            pass
    return len(known_names)

# ── Camera thread ─────────────────────────────────────────────────
def camera_thread():
    global session_state

    camera = cv2.VideoCapture(CAMERA_INDEX)
    if not camera.isOpened():
        socketio.emit("error", {"msg": "Cannot open camera"})
        session_state["running"] = False
        return

    session_state["next_snap_due"] = time.time()
    frame_count = 0
    last_locations, last_names_det, last_rolls_det, last_confs = [], [], [], []

    while session_state["running"]:
        ok, frame = camera.read()
        if not ok:
            break

        frame_count += 1

        # Process face detection every 3 frames
        if frame_count % 3 == 0:
            small = cv2.resize(frame, (0,0), fx=0.25, fy=0.25)
            rgb   = cv2.cvtColor(small, cv2.COLOR_BGR2RGB)
            locs  = face_recognition.face_locations(rgb)
            encs  = face_recognition.face_encodings(rgb, locs)
            last_locations   = locs
            last_names_det   = []
            last_rolls_det   = []
            last_confs       = []
            for enc in encs:
                matches   = face_recognition.compare_faces(known_encodings, enc, tolerance=TOLERANCE)
                distances = face_recognition.face_distance(known_encodings, enc)
                name, roll, conf = "Unknown", "", 0
                if True in matches:
                    best = np.argmin(distances)
                    if matches[best]:
                        name = known_names[best]
                        roll = known_rolls[best]
                        conf = round((1 - distances[best]) * 100, 1)
                last_names_det.append(name)
                last_rolls_det.append(roll)
                last_confs.append(conf)

        # Draw boxes on display frame
        display = frame.copy()
        for (top,right,bottom,left), name, roll, conf in zip(
                last_locations, last_names_det, last_rolls_det, last_confs):
            top*=4; right*=4; bottom*=4; left*=4
            color = (0,200,0) if name != "Unknown" else (0,0,220)
            cv2.rectangle(display, (left,top), (right,bottom), color, 2)
            cv2.rectangle(display, (left,bottom), (right,bottom+45), color, cv2.FILLED)
            cv2.putText(display, f"{roll} - {name}" if name!="Unknown" else "Unknown",
                        (left+6, bottom+16), cv2.FONT_HERSHEY_DUPLEX, 0.55, (255,255,255), 1)
            cv2.putText(display, f"Match: {conf}%" if name!="Unknown" else "Not in database",
                        (left+6, bottom+36), cv2.FONT_HERSHEY_DUPLEX, 0.45, (255,255,255), 1)

        # Countdown overlay
        time_remaining = max(0, session_state["next_snap_due"] - time.time())
        m = int(time_remaining)//60; s = int(time_remaining)%60
        cv2.rectangle(display, (0,0), (640,36), (30,30,30), cv2.FILLED)
        cv2.putText(display,
                    f"Snap {session_state['snap_count']}/{session_state['total_snaps']}  |  Next snap in: {m}m {s}s",
                    (10,24), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255,255,255), 1)

        # Encode frame as JPEG and send to browser via SocketIO
        _, buffer = cv2.imencode(".jpg", display, [cv2.IMWRITE_JPEG_QUALITY, 60])
        frame_b64 = base64.b64encode(buffer).decode("utf-8")
        socketio.emit("frame", {
            "img"          : frame_b64,
            "snap_count"   : session_state["snap_count"],
            "total_snaps"  : session_state["total_snaps"],
            "time_remaining": f"{m}m {s}s",
            "detected"     : [n for n in last_names_det if n != "Unknown"]
        })

        # Take snapshot at interval
        now = time.time()
        if now >= session_state["next_snap_due"] and session_state["snap_count"] < session_state["total_snaps"]:
            session_state["snap_count"] += 1
            session_state["next_snap_due"] = now + session_state["interval"] * 60

            # Save snapshot
            snap_file = os.path.join(SNAPSHOTS_FOLDER,
                f"snap{session_state['snap_count']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg")
            cv2.imwrite(snap_file, frame)

            # Record who was found
            found = [n for n in last_names_det if n != "Unknown"]
            for name in session_state["tracker"]:
                session_state["tracker"][name]["snaps"].append(name in found)

            log_entry = f"Snap {session_state['snap_count']} at {datetime.now().strftime('%H:%M:%S')} — detected: {', '.join(found) if found else 'nobody'}"
            session_state["log"].append(log_entry)
            socketio.emit("snap_taken", {"log": log_entry, "snap_count": session_state["snap_count"]})

            # If all snaps done — finish
            if session_state["snap_count"] >= session_state["total_snaps"]:
                finish_session()
                break

        time.sleep(0.03)

    camera.release()
    session_state["running"] = False

def finish_session():
    global session_state
    session_state["running"] = False

    attendance_data = []
    for name, data in session_state["tracker"].items():
        snaps_present = sum(data["snaps"])
        verdict = "Present" if snaps_present >= session_state["min_snaps"] else "Absent"
        attendance_data.append({
            "name": name, "roll": data["roll"],
            "snaps": data["snaps"], "verdict": verdict
        })

    date_str  = datetime.now().strftime("%d-%b-%Y")
    subject   = session_state["subject"]
    filepath  = generate_excel(attendance_data, subject, date_str)
    total     = len(attendance_data)
    present   = sum(1 for s in attendance_data if s["verdict"] == "Present")
    absent    = total - present
    pct       = round(present/total*100, 1) if total > 0 else 0

    send_email(filepath, subject, date_str, present, absent, total, pct)

    record = {
        "date": date_str, "subject": subject,
        "present": present, "absent": absent,
        "total": total, "pct": pct,
        "students": attendance_data
    }
    save_history(record)
    socketio.emit("session_done", record)

def generate_excel(attendance_data, subject, date_str):
    filename = f"attendance_{subject.replace(' ','_')}_{date_str}.xlsx"
    filepath = os.path.join(REPORTS_FOLDER, filename)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"
    hf = PatternFill("solid", fgColor="1F4E79")
    pf = PatternFill("solid", fgColor="C6EFCE")
    af = PatternFill("solid", fgColor="FFC7CE")
    headers = ["#","Roll No","Student Name","Date","Snap 1","Snap 2","Snap 3","Snap 4","Snap 5","Present Snaps","Total Snaps","Verdict"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hf
        cell.alignment = Alignment(horizontal="center")
    for i, s in enumerate(attendance_data, 1):
        syms = ["Y" if x else "N" for x in s["snaps"]]
        while len(syms) < 5: syms.append("-")
        ws.append([i, s["roll"], s["name"], date_str,
                   syms[0],syms[1],syms[2],syms[3],syms[4],
                   sum(s["snaps"]), len(s["snaps"]), s["verdict"]])
        vcell = ws.cell(i+1, 12)
        vcell.fill = pf if s["verdict"]=="Present" else af
        vcell.font = Font(bold=True, color="276221" if s["verdict"]=="Present" else "9C0006")
    total   = len(attendance_data)
    present = sum(1 for s in attendance_data if s["verdict"]=="Present")
    pct     = round(present/total*100,1) if total>0 else 0
    ws.append([])
    for row in [["","","Total","","","","","","","","",total],
                ["","","Present","","","","","","","","",present],
                ["","","Absent","","","","","","","","",total-present],
                ["","","Attendance %","","","","","","","","",f"{pct}%"]]:
        ws.append(row)
    for col,w in {"A":4,"B":10,"C":22,"D":14,"E":8,"F":8,"G":8,"H":8,"I":8,"J":14,"K":12,"L":12}.items():
        ws.column_dimensions[col].width = w
    wb.save(filepath)
    return filepath

def send_email(filepath, subject, date_str, present, absent, total, pct):
    msg = MIMEMultipart()
    msg["From"]    = SENDER_EMAIL
    msg["To"]      = FACULTY_EMAIL
    msg["Subject"] = f"Attendance Report - {subject} - {date_str}"
    body = f"""Dear Faculty,

Attendance report for today's class is attached.

Subject    : {subject}
Date       : {date_str}
Present    : {present} / {total}
Absent     : {absent} / {total}
Attendance : {pct}%

Regards,
Automated Face Recognition Attendance System"""
    msg.attach(MIMEText(body, "plain"))
    with open(filepath, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(filepath)}")
    msg.attach(part)
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            server.sendmail(SENDER_EMAIL, FACULTY_EMAIL, msg.as_string())
    except Exception as e:
        print(f"Email failed: {e}")

# ══════════════════════════════════════════════════════════════════
# ROUTES
# ══════════════════════════════════════════════════════════════════

@app.route("/")
def dashboard():
    count   = enroll_students()
    history = load_history()
    today   = [h for h in history if h["date"] == datetime.now().strftime("%d-%b-%Y")]
    return render_template("dashboard.html",
                           student_count=count,
                           history=history[-5:][::-1],
                           today=today)

@app.route("/start", methods=["POST"])
def start_session():
    global session_state
    if session_state["running"]:
        return jsonify({"error": "Session already running"})

    subject  = request.form.get("subject", "General")
    duration = int(request.form.get("duration", 50))
    interval = int(request.form.get("interval", 10))
    min_snaps= int(request.form.get("min_snaps", 4))

    enroll_students()
    if not known_names:
        return jsonify({"error": "No students enrolled"})

    session_state.update({
        "running"     : True,
        "subject"     : subject,
        "duration"    : duration,
        "interval"    : interval,
        "min_snaps"   : min_snaps,
        "snap_count"  : 0,
        "total_snaps" : duration // interval,
        "tracker"     : {n: {"roll": r, "snaps": []} for n,r in zip(known_names, known_rolls)},
        "log"         : [],
        "start_time"  : time.time(),
        "next_snap_due": time.time(),
    })

    thread = threading.Thread(target=camera_thread)
    thread.daemon = True
    thread.start()

    return jsonify({"ok": True})

@app.route("/stop", methods=["POST"])
def stop_session():
    session_state["running"] = False
    return jsonify({"ok": True})

@app.route("/live")
def live():
    return render_template("live.html",
                           subject=session_state["subject"],
                           running=session_state["running"])

@app.route("/students")
def students():
    files = [f for f in os.listdir(STUDENTS_FOLDER)
             if f.lower().endswith((".jpg",".jpeg",".png"))]
    student_list = []
    for f in files:
        base  = os.path.splitext(f)[0]
        parts = base.split("_")
        roll  = parts[0] if parts else ""
        name  = " ".join(parts[1:]).replace("-"," ").title() if len(parts)>1 else f
        student_list.append({"roll": roll, "name": name, "file": f})
    return render_template("students.html", students=student_list)

@app.route("/upload_student", methods=["POST"])
def upload_student():
    roll  = request.form.get("roll","").strip()
    name  = request.form.get("name","").strip().replace(" ","_")
    photo = request.files.get("photo")
    if not roll or not name or not photo:
        return redirect(url_for("students"))
    filename = f"{roll}_{name}.jpg"
    photo.save(os.path.join(STUDENTS_FOLDER, filename))
    enroll_students()
    return redirect(url_for("students"))

@app.route("/delete_student/<filename>")
def delete_student(filename):
    path = os.path.join(STUDENTS_FOLDER, filename)
    if os.path.exists(path):
        os.remove(path)
    return redirect(url_for("students"))

@app.route("/reports")
def reports():
    history = load_history()
    files   = os.listdir(REPORTS_FOLDER)
    return render_template("reports.html",
                           history=history[::-1],
                           files=files)

@app.route("/download/<filename>")
def download(filename):
    return send_file(os.path.join(REPORTS_FOLDER, filename), as_attachment=True)

@app.route("/status")
def status():
    return jsonify(session_state)

if __name__ == "__main__":
    enroll_students()
    print("\n" + "="*50)
    print("  Attendance Web App running!")
    print("  Open browser and go to: http://localhost:5000")
    print("="*50 + "\n")
    socketio.run(app, host="0.0.0.0", port=5000, debug=False)

